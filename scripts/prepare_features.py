"""
prepare_features.py
Loads the raw ecoinvent LCIA sheet, builds proper readable column names
from the Method/Category/Indicator header rows, and saves a clean CSV
for reuse in clustering and anomaly detection (avoids re-reading the
slow 20MB Excel file every time).
"""

import openpyxl
import pandas as pd
from pathlib import Path

SOURCE_FILE = Path(__file__).parent.parent / "data" / "RAJ_DISS.xlsx"
OUTPUT_FILE = Path(__file__).parent.parent / "data" / "clean_features.csv"


def build_column_names():
    wb = openpyxl.load_workbook(SOURCE_FILE, read_only=True, data_only=True)
    ws = wb["LCIA"]
    rows = list(ws.iter_rows(min_row=1, max_row=4, values_only=True))
    methods, cats, inds, units = rows[0], rows[1], rows[2], rows[3]

    names = []
    for i in range(len(methods)):
        if i < 7:
            continue  # metadata columns handled separately
        m, c, ind, u = methods[i], cats[i], inds[i], units[i]
        if m is None:
            continue  # skip the 3 blank trailing columns
        # Build a short, readable, unique name
        clean_name = f"{c}: {ind} [{m}]"
        names.append((i, clean_name))
    return names


def main():
    print("Reading header structure...")
    indicator_cols = build_column_names()
    print(f"Found {len(indicator_cols)} real indicator columns")

    print("Loading full data (this takes a minute)...")
    wb = openpyxl.load_workbook(SOURCE_FILE, read_only=True, data_only=True)
    ws = wb["LCIA"]

    records = []
    for row in ws.iter_rows(min_row=5, values_only=True):
        if row[2] is None:  # skip empty rows
            continue
        record = {
            "activity_name": row[2],
            "geography": row[3],
            "reference_product_name": row[4],
            "reference_product_unit": row[5],
            "reference_product_amount": row[6],
        }
        for idx, name in indicator_cols:
            record[name] = row[idx]
        records.append(record)

    df = pd.DataFrame(records)
    print(f"Final shape: {df.shape}")

    df.to_csv(OUTPUT_FILE, index=False)
    print(f"Saved clean feature file to {OUTPUT_FILE}")


if __name__ == "__main__":
    main()